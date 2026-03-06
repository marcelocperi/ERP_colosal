from openpyxl import load_workbook
import sys

# Force UTF-8 output
if sys.stdout.encoding.lower() != 'utf-8':
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except:
        pass

file_path = 'libros_enriquecidos_Empresa_1.xlsx'
wb = load_workbook(file_path, read_only=True)
sheet = wb.active

# Get headers
headers = [cell.value for cell in sheet[1]]
print(f"Headers: {headers}")

ebook_idx = -1
if 'Ebook' in headers:
    ebook_idx = headers.index('Ebook')

counts = {}
total = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    total += 1
    if ebook_idx != -1:
        val = str(row[ebook_idx])
        counts[val] = counts.get(val, 0) + 1

print(f"Total rows: {total}")
print(f"Ebook counts: {counts}")
