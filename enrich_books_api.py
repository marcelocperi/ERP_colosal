import mariadb
import time
import json
import traceback
import sys
import os
from datetime import datetime
from database import DB_CONFIG, get_db_pool
from services import library_api_service
from core.security_utils import sanitize_filename
from services.enrichment.strategies import get_books_to_process_query
from services.enrichment.efficiency import EfficiencyManager

# Configurar logging con caracteres seguros para terminal
from core.logger_config import configure_windows_console, setup_logger

configure_windows_console()

# Logger principal de este proceso
logger = setup_logger('enrich_books', log_file='enrich_books.log')

# Directorio de scripts (necesario para excels)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

from core.concurrency import should_stop, update_heartbeat, unregister_thread, register_thread, get_active_tasks


from services.enrichment.processor import BookEnrichmentProcessor

async def process_books_batch(limit=100, deep_scan=False, enterprise_id=1, strategy='conservative', global_processed=0):
    conn = None
    try:
        conn = get_db_pool().get_connection()
        cursor = conn.cursor(dictionary=True)
        efficiency_mgr = EfficiencyManager(conn)
        processor = BookEnrichmentProcessor(conn, efficiency_mgr)
        
        # 1. Obtener libros según estrategia
        query, params = get_books_to_process_query(strategy, enterprise_id, limit, deep_scan=deep_scan)
        cursor.execute(query, params)
        libros = cursor.fetchall()
        
        if not libros:
            return 0

        # --- OPTIMIZACIÓN N+1 ---
        # Pre-cargar qué libros ya tienen archivos digitales
        book_ids = [lib['id'] for lib in libros]
        files_map = processor.get_existing_files_map(book_ids, enterprise_id=enterprise_id)
        
        # Obtener ranking de eficiencia una sola vez por lote
        db_ranking = efficiency_mgr.get_service_ranking()
        
        updated = 0
        task_id = f"enrich_{enterprise_id}"
        
        for i, lib in enumerate(libros):
            # Heartbeat & Stop signals
            if i % 5 == 0: # Cada 5 libros para balancear performance
                if await should_stop(task_id):
                    logger.info("Stop signal detected. Terminating batch...")
                    break
                await update_heartbeat(task_id)

            libro_nombre = lib.get('nombre', 'Sin título')
            logger.info(f"[{i+1}/{len(libros)}] Procesando: {libro_nombre}")

            # UI Update
            try:
                cursor.execute("INSERT INTO system_stats (key_name, enterprise_id, value_str) VALUES ('batch_status', %s, %s) ON DUPLICATE KEY UPDATE value_str = %s", ('batch_status', enterprise_id, f"Procesando: {libro_nombre[:40]}...", f"Procesando: {libro_nombre[:40]}..."))
                cursor.execute("INSERT INTO system_stats (key_name, enterprise_id, value_int) VALUES ('batch_processed', %s, %s) ON DUPLICATE KEY UPDATE value_int = %s", ('batch_processed', enterprise_id, global_processed + i + 1, global_processed + i + 1))
                conn.commit()
            except: pass

            # Lógica de enriquecimiento delegada al procesador
            has_file = lib['id'] in files_map
            success, api_data = processor.enrich_book(lib, enterprise_id, db_ranking, deep_scan, has_file)

            # Persistencia de resultados
            current_meta = lib.get('metadata_json')
            metadata = json.loads(current_meta) if current_meta else {}
            
            # Normalizar booleanos de strings (legacy fix)
            for b_field in ['archivo_local', 'con_portada', 'con_descripcion']:
                if metadata.get(b_field) == 'true': metadata[b_field] = True
                elif metadata.get(b_field) == 'false': metadata[b_field] = False

            if success or api_data.get('archivo_id'):
                processor.update_book_record(lib, metadata, api_data)
                updated += 1
                logger.info(f"  [OK] Registro de {lib['id']} actualizado.")
            else:
                # Marcar como revisado en el JSON aunque no haya cambios para que no re-procese
                metadata['api_checked'] = 2
                cursor.execute("UPDATE stk_articulos SET metadata_json = %s WHERE id = %s", (json.dumps(metadata), lib['id'],))
                logger.warning(f"  [!] Sin datos nuevos para {lib['id']}.")

            conn.commit()
            time.sleep(0.2)
        
        # Ciclo de aprendizaje
        await efficiency_mgr.rotate_learning_cycle()
        
        return len(libros)
    except Exception as e:
        logger.error(f"ERROR CRÍTICO en el lote: {str(e)}")
        logger.error(traceback.format_exc())
        return 0
    finally:
        if conn: conn.close()

def export_to_excel(enterprise_id=1):
    """Exporta los libros completados a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        logger.info(f"Generando reporte Excel para Empresa {enterprise_id}...")
        
        conn = get_db_pool().get_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                l.id,
                l.nombre,
                l.codigo as isbn,
                l.modelo as autor,
                l.marca as editorial,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.lengua')) as lengua,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.origen')) as origen,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.descripcion')) as descripcion,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.cover_url')) as cover_url,
                JSON_EXTRACT(l.metadata_json, '$.paginas') as paginas,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.genero')) as genero,
                JSON_UNQUOTE(JSON_EXTRACT(l.metadata_json, '$.ebook_url')) as ebook_url,
                JSON_EXTRACT(l.metadata_json, '$.archivo_local') as tiene_archivo
            FROM stk_articulos l
            WHERE l.enterprise_id = %s AND JSON_EXTRACT(l.metadata_json, '$.api_checked') = 2
            ORDER BY l.nombre
        """, (enterprise_id,))
        
        libros = cursor.fetchall()
        conn.close()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libros Enriquecidos"
        
        # Encabezados
        headers = ['ID', 'Título', 'ISBN', 'Autor', 'Editorial', 'Lengua', 'Origen', 'Páginas', 'Género', 'Desc.', 'Ebook', 'Portada']
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos
        for libro in libros:
            # Determinar estado del ebook
            ebook_status = "No"
            # Evaluar json extraction que puede venir como string 'true'/true o 1 desde MariaDB
            archivo_ok = libro['tiene_archivo'] in [True, 1, 'true', '1']
            
            if archivo_ok:
                ebook_status = "Local (PDF/EPUB)"
            elif libro['ebook_url']:
                ebook_status = "Link (Preview/Web)"
                
            ws.append([
                libro['id'],
                libro['nombre'],
                libro['isbn'],
                libro['autor'],
                libro['editorial'],
                libro['lengua'],
                libro['origen'],
                libro['paginas'],
                libro['genero'],
                'Sí' if libro['descripcion'] else 'No',
                ebook_status,
                'Sí' if libro['cover_url'] else 'No'
            ])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 10
        
        # Guardar
        # Guardar (Nombre fijo para actualización constante + Backup con timestamp)
        main_filename = os.path.join(SCRIPT_DIR, f"libros_enriquecidos_Empresa_{enterprise_id}.xlsx")
        timestamp_filename = os.path.join(SCRIPT_DIR, f"libros_enriquecidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        wb.save(main_filename)
        # wb.save(timestamp_filename) # Comentado para no saturar de archivos
        
        logger.info(f"✓ Reporte Excel actualizado: {main_filename}")
        logger.info(f"  Total de libros: {len(libros)}")
        
        return main_filename
        
    except ImportError:
        logger.warning("⚠ openpyxl no está instalado. Instala con: pip install openpyxl")
        return None
    except Exception as e:
        logger.error(f"Error generando Excel: {str(e)}")
        logger.error(traceback.format_exc())
        return None

async def process_until_finished(batch_size=100, deep_scan=False, enterprise_id=1, strategy='conservative'):
    scan_type = f"PROFUNDO ({strategy.upper()})" if deep_scan else "NORMAL"
    logger.info(f"{'='*60}")
    logger.info(f"Iniciando proceso {scan_type} de enriquecimiento para Empresa ID: {enterprise_id}...")
    logger.info(f"{'='*60}")
    
    # Resetear contador al inicio
    try:
        conn = get_db_pool().get_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO system_stats (key_name, enterprise_id, value_int) VALUES ('batch_processed', %s, 0) ON DUPLICATE KEY UPDATE value_int = 0", (enterprise_id,))
        cursor.execute("INSERT INTO system_stats (key_name, enterprise_id, value_str) VALUES ('batch_status', %s, %s) ON DUPLICATE KEY UPDATE value_str = %s", (enterprise_id, f"Iniciando {scan_type}...", f"Iniciando {scan_type}..."))
        conn.commit()
        conn.close()
        logger.info("Contador y estado reseteados")
    except Exception as e:
        logger.warning(f"No se pudo resetear contador: {e}")

    from core.concurrency import should_stop, unregister_thread, register_thread, get_active_tasks
    
    # Check if already running (with stale protection)
    active_tasks = await get_active_tasks(enterprise_id=enterprise_id)
    task_id = f"enrich_{enterprise_id}"
    if task_id in active_tasks:
        task_info = active_tasks[task_id]
        # If it's stopping, or hasn't updated heartbeat in 2 minutes, we allow a takeover
        # Wait, if it's the same PID, it's definitely us.
        if task_info.get('status') == 'STOPPING':
            logger.info(f"⚠ La tarea {task_id} está en estado STOPPING. Forzando limpieza para iniciar nueva.")
            await unregister_thread(task_id)
        else:
            logger.warning(f"!!! Task {task_id} is already running. Aborting count.")
            return

    total_processed = 0
    batch_num = 1
    await register_thread(
        f"enrich_{enterprise_id}", 
        f"Catálogo completo (Ent: {enterprise_id})",
        process_name=f"Enriquecimiento {scan_type}",
        priority=2, # Alta prioridad
        source_origin='CLI',
        enterprise_id=enterprise_id
    )
    
    try:
        while True:
            if await should_stop(f"enrich_{enterprise_id}"):
                logger.info(f"!!! STOP SIGNAL RECEIVED for Enterprise {enterprise_id}. Terminating...")
                break
            
            from core.concurrency import update_heartbeat
            await update_heartbeat(f"enrich_{enterprise_id}", status="RUNNING")
                
            logger.info(f"Intentando procesar lote #{batch_num}...")
            processed = await process_books_batch(batch_size, deep_scan, enterprise_id, strategy=strategy, global_processed=total_processed)
            if processed == 0:
                logger.info("No hay más libros pendientes")
                break
            total_processed += processed
            logger.info(f"Acumulado: {total_processed} libros procesados")
            
            # Exportar a Excel cada lote para visibilidad del usuario
            if total_processed > 0:
                logger.info("  [+] Actualizando reporte Excel incremental...")
                export_to_excel(enterprise_id)
                
            batch_num += 1
            time.sleep(2)
    finally:
        await unregister_thread(f"enrich_{enterprise_id}")
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Proceso finalizado. Total procesados: {total_processed} libros")
    logger.info(f"{'='*60}")
    
    try:
        conn = get_db_pool().get_connection()
        cursor = conn.cursor()
        # No resetear a 0 el conteo para que el usuario vea el resultado final acumulado en la UI
        cursor.execute("INSERT INTO system_stats (key_name, enterprise_id, value_str) VALUES ('batch_status', %s, %s) ON DUPLICATE KEY UPDATE value_str = %s", ('batch_status', enterprise_id, f"Finalizado (Total: {total_processed})", f"Finalizado (Total: {total_processed})"))
        conn.commit()
        conn.close()
    except: pass
    
    # Generar Excel al finalizar
    # Generar Excel al finalizar si se procesaron libros
    if total_processed > 0:
        logger.info("\nGenerando reporte Excel...")
        export_to_excel(enterprise_id)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--enterprise", type=int, default=1)
    parser.add_argument("--deep", action="store_true")
    parser.add_argument("--strategy", type=str, default="conservative")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--export", action="store_true")
    
    args = parser.parse_args()
    
    if args.export:
        export_to_excel(args.enterprise)
    else:
        import asyncio
        asyncio.run(process_until_finished(
            batch_size=args.limit, 
            deep_scan=args.deep, 
            enterprise_id=args.enterprise, 
            strategy=args.strategy
        ))

