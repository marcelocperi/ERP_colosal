# -*- coding: utf-8 -*-
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import secrets
import string

# Add module path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../multiMCP'))
from database import get_db_cursor
from services import email_service

OUTBOX_DIR = Path(__file__).parent.parent.parent / "outbox"
OUTBOX_DIR.mkdir(exist_ok=True)

class PurchaseOrderMailer:
    def __init__(self, enterprise_id):
        self.enterprise_id = enterprise_id
        
    def generate_security_hash(self, provider_id):
        """Genera hash único para PO: PO_PROV_70chars"""
        chars = string.ascii_letters + string.digits
        random_str = ''.join(secrets.choice(chars) for _ in range(70))
        return f"PO_PROV{provider_id}_{random_str}"

    async def create_order_from_quotation(self, cotizacion_id, existing_cursor=None):
        """Crea la Orden de Compra en BD basada en la Cotización Aprobada."""
        # Si ya viene un cursor de una transacción atómica superior, lo usamos.
        if existing_cursor:
            return await self._logic_create_order(cotizacion_id, existing_cursor)
        
        # Si no, creamos uno nuevo (comportamiento legacy)
        async with get_db_cursor(dictionary=True) as c:
            return await self._logic_create_order(cotizacion_id, c)

    async def _logic_create_order(self, cotizacion_id, c):
        """Lógica interna de creación de PO."""
        # Get Quotation & Items
        await c.execute("SELECT * FROM cmp_cotizaciones WHERE id = %s AND enterprise_id = %s", (cotizacion_id, self.enterprise_id))
        cot = await c.fetchone()
        if not cot: return None, "Cotización no encontrada"
        
        await c.execute("""
            SELECT i.*, a.nombre as articulo_nombre, a.codigo as articulo_codigo 
            FROM cmp_items_cotizacion i 
            JOIN stk_articulos a ON i.articulo_id = a.id 
            WHERE i.cotizacion_id = %s AND i.enterprise_id = %s
        """, (cotizacion_id, self.enterprise_id))
        items = await c.fetchall()
        
        # Create PO Header
        po_hash = await self.generate_security_hash(cot['proveedor_id'])
        
        # Insert PO
        await c.execute("""
            INSERT INTO cmp_ordenes_compra 
            (enterprise_id, proveedor_id, estado, fecha_emision, cotizacion_id, security_hash, total_estimado)
            VALUES (%s, %s, 'PENDIENTE_APROBACION_COMPRAS', NOW(), %s, %s, 0)
        """, (self.enterprise_id, cot['proveedor_id'], cotizacion_id, po_hash))
        po_id = c.lastrowid
        
        # Copy Items to PO Details
        total_estimado = 0
        po_items = []
        for item in items:
            cant_final = item.get('cantidad_ofrecida')
            if cant_final is None:
                cant_final = item['cantidad']
            
            precio = item.get('precio_cotizado') or 0
            
            if float(precio) <= 0 or float(cant_final) <= 0:
                continue

            subtotal = float(precio) * float(cant_final)
            total_estimado += subtotal
            
            await c.execute("""
                INSERT INTO cmp_detalles_orden 
                (enterprise_id, orden_id, articulo_id, cantidad, precio_unitario)
                VALUES (%s, %s, %s, %s, %s)
            """, (self.enterprise_id, po_id, item['articulo_id'], cant_final, precio))
            
            po_items.append({
                'articulo_id': item['articulo_id'],
                'codigo': item['articulo_codigo'],
                'descripcion': item['articulo_nombre'],
                'cantidad': cant_final,
                'precio': precio,
                'plazo': item.get('disponibilidad_dias', 0)
            })

        # Update PO Total
        await c.execute("UPDATE cmp_ordenes_compra SET total_estimado = %s WHERE id = %s AND enterprise_id = %s", (total_estimado, po_id, self.enterprise_id))

        return {'po_id': po_id, 'hash': po_hash, 'items': po_items, 'proveedor_id': cot['proveedor_id']}, None

    async def generate_excel_po(self, po_id, proveedor_nombre, items, po_hash):
        """Genera Excel de Orden de Compra (PO)."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"PO-{po_id}"
        
        # Hash en A1
        ws.merge_cells('A1:F1')
        cell_ref = ws.cell(row=1, column=1, value=f"PO REF: {po_hash}")
        cell_ref.font = Font(b=True, color="999999", size=8)
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="27AE60") # Green for PO
        
        headers = ["CODIGO", "DESCRIPCION", "CANTIDAD", "PRECIO UNITARIO", "SUBTOTAL", "PLAZO (Dias)"]
        
        ws.append([])
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        start_row = 3
        for i, item in enumerate(items):
            row = start_row + i
            ws.cell(row=row, column=1, value=item['codigo'])
            ws.cell(row=row, column=2, value=item['descripcion'])
            ws.cell(row=row, column=3, value=item['cantidad'])
            ws.cell(row=row, column=4, value=item['precio']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=float(item['cantidad'])*float(item['precio'])).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=item['plazo'])

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        filename = f"PO_{po_id}_{po_hash[:10]}.xlsx"
        filepath = OUTBOX_DIR / filename
        await wb.save(filepath)
        return str(filepath)

    def generate_html_body(self, empresa_nombre, po_id, po_hash):
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; background-color: #f7fff7; color: #333; }}
                .container {{ max-width: 600px; margin: 20px auto; background: #fff; padding: 40px; border-radius: 8px; border-top: 5px solid #27ae60; box-shadow: 0 4px 10px rgba(0,0,0,0.1); }}
                h2 {{ color: #27ae60; }}
                .hash-box {{ background: #eee; padding: 10px; font-family: monospace; font-size: 0.8rem; word-break: break-all; margin: 20px 0; }}
                .action-box {{ background: #e8f5e9; padding: 15px; border-radius: 5px; margin-top: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Purchase Order (PO) #{po_id}</h2>
                <p>Estimado Proveedor,</p>
                <p>Confirmamos la compra de los items detallados en el archivo adjunto, según su cotización previa.</p>
                
                <div class="action-box">
                    <strong>ACCIÓN REQUERIDA:</strong>
                    <p>Por favor responda a este correo con la palabra <strong>"SI"</strong> o <strong>"CONFIRMO"</strong> para validar y procesar esta Orden de Compra.</p>
                </div>

                <p>Código de Seguridad PO:</p>
                <div class="hash-box">{po_hash}</div>
                
                <p>Atentamente,<br><strong>{empresa_nombre}</strong></p>
            </div>
        </body>
        </html>
        """

    async def send_po_email(self, to_email, po_id, po_hash, excel_path, empresa_nombre=""):
        """Envía el email de Orden de Compra usando el servicio centralizado."""
        subject = f"Purchase Order #{po_id} - REF: {po_hash[:10]} - CONFIRMACION REQUERIDA"
        body = self.generate_html_body(empresa_nombre or "Nuestra Empresa", po_id, po_hash)
        
        # El servicio _enviar_email ya maneja adjuntos como paths
        success, error = await email_service._enviar_email(
            recipient_email=to_email,
            subject=subject,
            html_content=body,
            attachments=[excel_path] if excel_path else None,
            enterprise_id=self.enterprise_id
        )
        
        if success:
            print(f"   [PO] Email PO #{po_id} enviado a {to_email}")
        return success, error

