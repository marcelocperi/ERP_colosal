# -*- coding: utf-8 -*-
"""
Quotation Mail Service (Real Sender)
=====================================
Envio real de correos con Excel adjunto usando Gmail SMTP.
"""
import sys, os, io, json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

# Add module path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../multiMCP'))
from database import get_db_cursor
from services import email_service  # Importar config de email

OUTBOX_DIR = Path(__file__).parent.parent.parent / "outbox"
OUTBOX_DIR.mkdir(exist_ok=True)

import secrets
import string

# ... imports ...

class QuotationMailer:
    def __init__(self, enterprise_id):
        self.enterprise_id = enterprise_id
        
    async def generate_security_hash(self, provider_code):
        """Genera hash único: CODIGO_70chars"""
        chars = string.ascii_letters + string.digits
        random_str = ''.join(secrets.choice(chars) for _ in range(70))
        # Limpiar provider code de caracteres raros
        safe_code = "".join([c for c in provider_code if c.isalnum()])
        return f"{safe_code}_{random_str}"

        # METADATA DE SEGURIDAD (Visible y Protegida)
        # La pondremos en la fila 1 para que el ERP la lea facil, pero bloqueada.
        ws.merge_cells('A1:B1')
        ws.cell(row=1, column=1, value="ID SOLICITUD:")
        ws.cell(row=1, column=3, value=cotizacion_id)
        
        ws.merge_cells('D1:E1')
        ws.cell(row=1, column=4, value="FECHA LÍMITE:")
        ws.cell(row=1, column=6, value=fecha_vencimiento if fecha_vencimiento else "N/A")

        ws.cell(row=1, column=8, value=f"HASH: {security_hash[:12]}") # Hash corto visual

        # Estilos Header Superior
        meta_font = Font(bold=True, size=10, color="2C3E50")
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = meta_font
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="left")

        # ESTILOS
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        editable_fill = PatternFill("solid", fgColor="FFFFCC")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = [
            "ID Articulo", "CODIGO", "DESCRIPCION", "CANT. A COTIZAR",
            "COSTO UNITARIO ($)", "FECHA EST. ENTREGA", "OBSERVACIONES"
        ]
        
        # Headers en fila 3 para dejar aire
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # VALIDACIONES
        dv_fecha = DataValidation(type="date", operator="greaterThanOrEqual", formula1="TODAY()", allow_blank=True)
        ws.add_data_validation(dv_fecha)
        
        dv_costo = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        ws.add_data_validation(dv_costo)

        # Protección de hoja - Password derivado del hash (solo el ERP lo sabe)
        password_erp = f"ERP_{security_hash[:8]}" 
        ws.protection.set_password(password_erp)
        ws.protection.sheet = True
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.insertColumns = True # Permitir insertar comentarios si quieren
        ws.protection.deleteColumns = True
        ws.protection.deleteRows = True

        unlocked = Protection(locked=False)
        locked = Protection(locked=True)

        start_row = 4
        for i, item in enumerate(items):
            row = start_row + i
            # Estática (Bloqueada explícitamente)
            c_id = ws.cell(row=row, column=1, value=item['articulo_id'])
            c_id.protection = locked
            
            c_cod = ws.cell(row=row, column=2, value=item['codigo_interno'])
            c_cod.protection = locked
            
            c_nom = ws.cell(row=row, column=3, value=item['nombre_articulo'])
            c_nom.protection = locked
            
            # Editable: Cantidad Ofrecida (Iniciamos con la solicitada)
            c_cant = ws.cell(row=row, column=4, value=item['cantidad'])
            c_cant.fill = editable_fill; c_cant.border = border
            c_cant.protection = unlocked
            
            # Editable: Fecha Est. Entrega
            c_fecha = ws.cell(row=row, column=6)
            c_fecha.fill = editable_fill; c_fecha.border = border
            c_fecha.number_format = 'DD/MM/YYYY'; c_fecha.protection = unlocked
            dv_fecha.add(c_fecha)

            # Editable: Observaciones
            c_obs = ws.cell(row=row, column=7)
            c_obs.fill = editable_fill; c_obs.border = border
            c_obs.protection = unlocked

        # Anchos
        widths = {'A':10, 'B':15, 'C':45, 'D':18, 'E':20, 'F':20, 'G':35}
        for k, v in widths.items(): ws.column_dimensions[k].width = v
        
        filename = f"SC_{cotizacion_id}_{security_hash[:10]}.xlsx"
        filepath = OUTBOX_DIR / filename
        await wb.save(filepath)
        return str(filepath)

    def generate_html_body(self, empresa_nombre, cotizacion_id, security_hash, fecha_vencimiento=None):
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; background-color: #f4f4f4; margin: 0; padding: 0; }}
                .container {{ max-width: 600px; margin: 20px auto; background: #ffffff; padding: 40px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                .header {{ border-bottom: 2px solid #3498db; padding-bottom: 20px; margin-bottom: 30px; }}
                h2 {{ color: #2c3e50; margin: 0; }}
                .cta-box {{ background-color: #e8f6fd; border-left: 5px solid #3498db; padding: 20px; margin: 30px 0; }}
                .hash-box {{ background-color: #f8f9fa; border: 1px dashed #ccc; padding: 10px; font-family: monospace; font-size: 0.8rem; color: #555; word-break: break-all; }}
                .footer {{ font-size: 12px; color: #999; margin-top: 40px; text-align: center; border-top: 1px solid #eee; padding-top: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>Solicitud de Cotización #{cotizacion_id}</h2>
                    <p>De: <strong>{empresa_nombre}</strong></p>
                </div>
                <div class="content">
                    <p>Estimado Proveedor,</p>
                    <p>Solicitamos cotización de los artículos adjuntos.</p>
                    <div class="cta-box">
                        <strong>Instrucciones:</strong>
                        <ol>
                            <li>Descargue el Excel adjunto.</li>
                            <li>Complete <strong>Costo Unitario</strong> y <strong>Disponibilidad</strong> (amarillo).</li>
                            <li>Si no cotiza, indique 0 o deje vacío.</li>
                            <li>Reenvíe este correo con el Excel completo.</li>
                        </ol>
                        <p style="color: #e74c3c; font-weight: bold; margin-top: 15px;">
                            <i class="fas fa-clock"></i> FECHA LÍMITE DE PRESENTACIÓN: {fecha_vencimiento if fecha_vencimiento else 'Consultar'}
                        </p>
                    </div>
                    <p>Código de Seguridad del Proceso:</p>
                    <div class="hash-box">{security_hash}</div>
                    <p>Por favor no modifique este código en el Excel para garantizar el procesamiento automático.</p>
                    <p>Atentamente,<br><strong>Dpto. Compras - {empresa_nombre}</strong></p>
                </div>
                <div class="footer"><p>ERP System - {empresa_nombre}</p></div>
            </div>
        </body>
        </html>
        """

    async def send_email_real(self, to_email, subject, body, attachment_path):
        """
        Envía correo real usando el servicio centralizado.
        Referencia: https://github.com/marcelocperi/BibliotecaWEBPy
        """
        success, error = await email_service._enviar_email(
            recipient_email=to_email,
            subject=subject,
            html_content=body,
            attachments=[attachment_path] if attachment_path else None,
            enterprise_id=self.enterprise_id
        )
        
        if success:
            print("   [OK] Email de cotización enviado exitosamente.")
        return success, error

    async def process_pending_quotations(self):
        processed_count = 0
        async with get_db_cursor(dictionary=True) as c:
            # Buscar cotizaciones ENVIADA que aun no tengan hash (retrocompatibilidad) o simplemente ENVIADA
            await c.execute("""
                SELECT c.id, c.proveedor_id, p.nombre as razon_social, p.cuit, p.email, e.nombre as empresa_nombre, c.security_hash, c.fecha_vencimiento
                FROM cmp_cotizaciones c
                JOIN erp_terceros p ON c.proveedor_id = p.id
                JOIN sys_enterprises e ON c.enterprise_id = e.id
                WHERE c.estado = 'ENVIADA' AND c.enterprise_id = %s
                ORDER BY c.id DESC LIMIT 5
            """, (self.enterprise_id,))
            
            pendientes = await c.fetchall()
            if not pendientes:
                print("No hay cotizaciones pendientes.")
                return 0
                
            print(f"Procesando {len(pendientes)} cotizaciones...")
            
            for cot in pendientes:
                # Generar Hash si no existe
                if not cot['security_hash']:
                    # Prefijo con Codigo Proveedor (o ID si no tiene codigo explicito)
                    # Usaremos ID_CUIT para ser unicos
                    prefijo = f"PROV{cot['proveedor_id']}"
                    current_hash = await self.generate_security_hash(prefijo)
                    
                    # Guardar en BD
                    await c.execute("UPDATE cmp_cotizaciones SET security_hash = %s WHERE id = %s AND enterprise_id = %s", (current_hash, cot['id'], self.enterprise_id))
                    cot['security_hash'] = current_hash # Actualizar dict local
                
                # Items
                await c.execute("""
                    SELECT i.articulo_id, a.codigo as codigo_interno, a.nombre as nombre_articulo, i.cantidad
                    FROM cmp_items_cotizacion i
                    JOIN stk_articulos a ON i.articulo_id = a.id
                    WHERE i.cotizacion_id = %s AND i.enterprise_id = %s
                """, (cot['id'], self.enterprise_id))
                items = await c.fetchall()
                
                # Generar Archivos con HASH y Protección
                excel_path = self.generate_excel_attachment(cot['id'], cot['razon_social'], items, cot['security_hash'], cot['fecha_vencimiento'])
                html_body = self.generate_html_body(cot['empresa_nombre'], cot['id'], cot['security_hash'], cot['fecha_vencimiento'])
                
                # HTML Log Local
                html_path = OUTBOX_DIR / f"Email_SC_{cot['id']}_{cot['security_hash'][:10]}.html"
                with open(html_path, "w", encoding="utf-8") as f: f.write(html_body)
                
                # ENVIAR CORREO REAL
                if not cot['email']:
                    print(f"   [SKIP] Saltando Cotización #{cot['id']} - El proveedor '{cot['razon_social']}' no tiene email.")
                    continue

                print(f"   Preparando envío a {cot['email']} (Hash: {cot['security_hash'][:15]}...)...")
                await self.send_email_real(
                    to_email=cot['email'], 
                    subject=f"Solicitud Cotización #{cot['id']} - REF: {cot['security_hash'][:10]}",
                    body=html_body,
                    attachment_path=excel_path
                )
                
                processed_count += 1
                
        return processed_count

if __name__ == "__main__":
    try:
        service = QuotationMailer(enterprise_id=0)
        await service.process_pending_quotations()
    except Exception as e:
        print(f"Error: {e}")
